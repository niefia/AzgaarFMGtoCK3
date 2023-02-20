import openpyxl
from openpyxl.styles import PatternFill
import random

# Open the Excel file and select the active worksheet
workbook = openpyxl.load_workbook('BFSoutput.xlsx')
worksheet = workbook.active

# Get the values from the 'nearest_town' column
values = [cell.value for cell in worksheet['C'][1:]]

# Find the unique values and assign a random color to each one
unique_values = list(set(values))
color_dict = {}
for value in unique_values:
    color_dict[value] = '{:06x}'.format(random.randint(0, 0xFFFFFF))

# Rename the column to 'color'
worksheet.cell(row=1, column=5).value = "color"

# Add a new column called 'color' and assign each row a color based on its 'nearest_town' value
for i, value in enumerate(values):
    cell = worksheet.cell(row=i+2, column=5)
    cell.value = color_dict[value]
    cell.fill = PatternFill(start_color=color_dict[value], end_color=color_dict[value], fill_type='solid')

# Save the updated worksheet to a new sheet called 'Output'
workbook.save('BFSoutput.xlsx')
