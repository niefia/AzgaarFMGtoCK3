import os
import json
import openpyxl
import zipfile
import subprocess

def modFile(modname, output_dir, modpath):
    escaped_output_dir = output_dir.replace("\\", "/")
    content = f"""version="1"
tags={{
        "Total Conversion"
}}
name="{modname}"
supported_version="1.3.1"
path="{escaped_output_dir}"
    """

    with open(os.path.join(modpath, f"{modname}.mod"), "w") as f:
        f.write(content)

    with open(os.path.join(output_dir, f"descriptor.mod"), "w") as f:
        f.write(content)

def biomeWrite(json_file_path, xlsx_file_path, image_folder_path):
    # Load the JSON data from the file
    with open(json_file_path, 'r') as f:
        data = json.load(f)

    # Extract the biome names and IDs from the "biomes" section of the data
    biomes = [(name, id) for name, id in zip(data['biomes']['name'], data['biomes']['i'])]

    # Create a new Excel workbook and worksheet to store the data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Biomes'

    # Write the biome names and IDs to the worksheet
    for row, (name, id) in enumerate(biomes, start=1):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=id)
        if name == 'Temperate deciduous forest':
            ws.cell(row=row, column=3, value='plains_01_noisy_mask')
        elif name == 'Grassland':
            ws.cell(row=row, column=3, value='steppe_01_mask')
        elif name == 'Tropical seasonal forest':
            ws.cell(row=row, column=3, value='forest_leaf_01_mask')
        elif name == 'Temperate rainforest':
            ws.cell(row=row, column=3, value='medi_grass_01_mask')
        elif name == 'Tropical rainforest':
            ws.cell(row=row, column=3, value='forest_jungle_01_mask')
        elif name == 'Wetland':
            ws.cell(row=row, column=3, value='wetlands_02_mask')
        elif name == 'Savanna':
            ws.cell(row=row, column=3, value='medi_lumpy_grass_mask')
        elif name == 'Taiga':
            ws.cell(row=row, column=3, value='forest_pine_01_mask')
        elif name == 'Hot desert':
            ws.cell(row=row, column=3, value='desert_01_mask')
        elif name == 'Cold desert':
            ws.cell(row=row, column=3, value='desert_02_mask')
        elif name == 'Tundra':
            ws.cell(row=row, column=3, value='forestfloor_mask')
        elif name == 'Glacier':
            ws.cell(row=row, column=3, value='snow_mask')
        elif name == 'Marine':
            ws.cell(row=row, column=3, value='beach_02_pebbles_mask')

    # Save the workbook to a file
    wb.save(xlsx_file_path)

    # Rename the biome image files to match the masks in the Excel worksheet
    for filename in os.listdir(image_folder_path):
        if filename.endswith('.png'):
            try:
                # Extract the number part of the filename (before the extension)
                number = filename[6:-4]
                number = int(number)
            except ValueError:
                # Skip files with invalid filenames
                continue

            # Find the corresponding row in the Excel worksheet
            for row in ws.iter_rows(min_row=1, max_col=3):
                if row[1].value == number:
                    # Generate the new filename for the source file
                    new_filename = f"{row[2].value}.png"

                    # Remove the existing file, if it exists
                    destination_file = os.path.join(image_folder_path, new_filename)
                    if os.path.exists(destination_file):
                        os.remove(destination_file)

                    # Rename the file
                    os.rename(os.path.join(image_folder_path, filename), destination_file)

def extract_zip_file(input_zip_file, output_folder):
    """
    Extracts the contents of a zip file to a specified output folder, skipping any file conflicts.

    Parameters:
    input_zip_file (str): The path to the input zip file.
    output_folder (str): The path to the output folder.

    Returns:
    None
    """
    with open(input_zip_file, 'rb') as f:
        zip_bytes = f.read()
    with open(os.path.join(output_folder, os.path.basename(input_zip_file)), 'wb') as f:
        f.write(zip_bytes)
    with zipfile.ZipFile(os.path.join(output_folder, os.path.basename(input_zip_file)), 'r') as zip_ref:
        for member in zip_ref.infolist():
            if os.path.exists(os.path.join(output_folder, member.filename)):
                print(f"Skipping {member.filename} (already exists in output folder)")
            else:
                zip_ref.extract(member, output_folder)


def modify_config(moddir, installdir, config_file_path):
    # Replace single backslashes with double backslashes in the paths
    moddir = moddir.replace("\\", "\\\\")
    installdir = installdir.replace("\\", "\\\\")

    with open(config_file_path, "r") as f:
        lines = f.readlines()

    for i, line in enumerate(lines):
        if line.startswith("moddirectory="):
            lines[i] = f"moddirectory={moddir}\n"
        elif line.startswith("installdir="):
            lines[i] = f"installdir={installdir}\n"
        elif line.startswith("removeVanilla=false"):
            lines[i] = "removeVanilla=true\n"

    with open(config_file_path, "w") as f:
        f.writelines(lines)


def run_jar(jar_path, cwd):

    print("Running Map Filler jar")
    # Set the current working directory to the directory where the script is located
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    # Command to run the jar file with the appropriate environment
    cmd = ["java", "-jar", jar_path]

    # Run the jar file in a subprocess with the appropriate environment and working directory
    subprocess.run(cmd, cwd=cwd)

