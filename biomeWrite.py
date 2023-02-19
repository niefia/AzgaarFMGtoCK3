import json
import openpyxl
import os

# Load the JSON data from the file
with open('noemoji.json', 'r') as f:
    data = json.load(f)

# Extract the biome names and IDs from the "biomes" section of the data
biomes = [(name, id) for name, id in zip(data['biomes']['name'], data['biomes']['i'])]

# Create a new Excel workbook and worksheet to store the data
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Biomes'

# Write the biome names and IDs to the worksheet
for row, (name, id) in enumerate(biomes, start=1):
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=id)
    if name == 'Temperate deciduous forest':
        ws.cell(row=row, column=3, value='plains_01_noisy_mask')
    elif name == 'Grassland':
        ws.cell(row=row, column=3, value='steppe_01_mask')
    elif name == 'Tropical seasonal forest':
        ws.cell(row=row, column=3, value='forest_leaf_01_mask')
    elif name == 'Temperate rainforest':
        ws.cell(row=row, column=3, value='medi_grass_01_mask')
    elif name == 'Tropical rainforest':
        ws.cell(row=row, column=3, value='forest_jungle_01_mask')
    elif name == 'Wetland':
        ws.cell(row=row, column=3, value='wetlands_02_mask')
    elif name == 'Savanna':
        ws.cell(row=row, column=3, value='medi_lumpy_grass_mask')
    elif name == 'Taiga':
        ws.cell(row=row, column=3, value='forest_pine_01_mask')
    elif name == 'Hot desert':
        ws.cell(row=row, column=3, value='desert_01_mask')
    elif name == 'Cold desert':
        ws.cell(row=row, column=3, value='desert_02_mask')
    elif name == 'Tundra':
        ws.cell(row=row, column=3, value='forestfloor_mask')
    elif name == 'Glacier':
        ws.cell(row=row, column=3, value='snow_mask')


# Save the workbook to a file
wb.save('biomes.xlsx')

# Rename the biome image files to match the masks in the Excel worksheet
for filename in os.listdir('terrain'):
    if filename.endswith('.png'):
        # Extract the number part of the filename (before the extension)
        number = filename[6:-4]
        print(number)
        # Find the corresponding row in the Excel worksheet
        for row in ws.iter_rows(min_row=2, max_col=3):
            if row[1].value == int(number)+1:
                # Rename the file to match the mask in column 3
                os.rename(os.path.join('terrain', filename), os.path.join('terrain', f"{row[2].value}.png"))
                break

