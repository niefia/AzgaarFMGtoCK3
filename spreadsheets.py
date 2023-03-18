import pandas as pd
import json
import random
import re
import hextorgb
import openpyxl

#removes emoji data from json, this is needed to avoid instability. previously jsonread.py
def remove_emoji_from_json(input_file_path, output_file_path):
    emoji_pattern = re.compile("["
        u"\U0001F600-\U0001F64F"  # emoticons
        u"\U0001F300-\U0001F5FF"  # symbols & pictographs
        u"\U0001F680-\U0001F6FF"  # transport & map symbols
        u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
        u"\U00002702-\U000027B0"
        u"\U000024C2-\U0001F251"
        "]+", flags=re.UNICODE)
    with open(input_file_path, 'r', encoding='utf-8') as json_file:
        data = json.load(json_file)
    for key, value in data.items():
        if isinstance(value, str):
            data[key] = emoji_pattern.sub(r'', value)
    with open(output_file_path, 'w') as json_file:
        json.dump(data, json_file)

#Adds unique colour to every barony. For cells generation method
def colorRandom(input_file, output_file):


    with open(input_file) as f:
        data = json.load(f)

    for feature in data['features']:
        feature['properties']['color'] = hextorgb.color_gen()

    with open(output_file, 'w') as f:
        json.dump(data, f)

#export the removed emoji json to spreadsheet, previously jsontoxlsprovicnes.py
def json_to_sheet(input_file_path, output_file_path):

    #Load the JSON data from the file
    with open(input_file_path) as file:
        data = json.load(file)

    # Extract the "states" and "provinces" fields from the data
    states = data["cells"]["states"]
    provinces = data["cells"]["provinces"]
    culture = data["cells"]["cultures"]
    religions = data["cells"]["religions"]
    burgs = data["cells"]["burgs"]

    # Create a list of dictionaries, where each dictionary represents a row of data for the states
    states_rows = []
    for cell in states:
        if "name" not in cell or "i" not in cell:
            continue  # skip this row if name or i is missing
        row = {
            "i": cell["i"],
            "name": cell["name"],
        }
        states_rows.append(row)

    # Create a list of dictionaries, where each dictionary represents a row of data for the provinces

    provinces_rows = []
    suffixes = ["castle", "town", "field", "pool","by","toft","worth","llyn","ay","y","ey","bost","caster","chester","cester","leigh","ley","borough","bury","burgh","wick"]
    names_set = set()  # to keep track of unique province names

    for cell in provinces:
        if not isinstance(cell, dict) or "name" not in cell:
            continue  # skip this row if cell is not a dictionary or name is missing

        name = cell["name"]
        base_name = name
        suffix_idx = 1
        while name in names_set:  # check if name is already in set
            name = base_name + suffixes[suffix_idx - 1]  # add a suffix to the name
            suffix_idx += 1
        names_set.add(name)  # add the unique name to the set

        row = {
            "i": cell["i"],
            "state": cell["state"],
            "center": cell["center"],
            "burg": cell["burg"],
            "name": name,
            "formName": cell["formName"],
            "fullName": cell["fullName"],
            "color": cell["color"],
        }
        provinces_rows.append(row)


    religions_rows = []
    for cell in religions:
        if isinstance(cell, dict):
            origins = cell.get("origins")
            origin = None
            if origins and isinstance(origins, list) and len(origins) > 0:
                origin = origins[0]
            row = {
                "i": cell.get("i"),
                "name": cell.get("name"),
                "type": cell.get("type"),
                "form": cell.get("form"),
                "deity": cell.get("deity"),
                "center": cell.get("center"),
                "origin": origin,
            }
            religions_rows.append(row)

    cultures_rows = []
    for cell in culture:
        if not isinstance(cell, dict) or "name" not in cell or "i" not in cell:
            continue  # skip this row if it's not a dictionary or name or i is missing
        name_words = cell["name"].split()
        row = {
            "i": cell["i"],
            "name": name_words[0],
        }
        if "type" in cell:
            row["type"] = cell["type"]
        if "origins" in cell and isinstance(cell["origins"], list) and len(cell["origins"]) > 0:
            row["origin"] = cell["origins"][0]
        cultures_rows.append(row)

    suffixes = ["castle", "town", "field", "pool"]
    names = set()
    burgs_rows = []



    suffixes = ["castle", "town", "field", "pool","by","toft","worth","llyn","ay","y","ey","bost","caster","chester","cester","leigh","ley","borough","bury","burgh","wick"]
    names = set()
    burgs_rows = []

    for cell in burgs:
        if isinstance(cell, dict) and cell:
            name = cell.get("name", None)
            suffix = ""
            while name + suffix in names:
                suffix = f"{random.choice(suffixes)}"
                if len(names) >= 4 * len(suffixes) * len(burgs):
                    print("ERROR: Could not generate unique names for all burgs.")
                    exit()
            names.add(name + suffix)
            row = {
                "i": cell.get("i", None),
                "cell": cell.get("cell", None),
                "name": name + suffix,
                "type": cell.get("type", None),
            }
            burgs_rows.append(row)


    burgs_df = pd.DataFrame(burgs_rows, columns=["i", "cell", "name"])

    # Create data frames from the lists of dictionaries
    states_df = pd.DataFrame(states_rows, columns=["i", "name"])
    provinces_df = pd.DataFrame(provinces_rows, columns=["i", "state", "center", "burg", "name", "formName", "fullName", "color"])
    cultures_df = pd.DataFrame(cultures_rows, columns=["i", "name","type","origin"])
    religion_df = pd.DataFrame(religions_rows, columns=["i", "name", "color", "culture", "type", "form", "deity", "center","origin"])
    burgs_df = pd.DataFrame(burgs_rows, columns=["i", "cell", "name","type"])

    # Save each data frame to a separate sheet in the same Excel file
    with pd.ExcelWriter(output_file_path) as writer:
        states_df.to_excel(writer, sheet_name="states", index=False)
        provinces_df.to_excel(writer, sheet_name="provinces", index=False)
        cultures_df.to_excel(writer, sheet_name="cultures", index=False)
        religion_df.to_excel(writer, sheet_name="religion", index=False)
        burgs_df.to_excel(writer, sheet_name="burgs", index=False)

#json_to_sheet('noemoji.json','ouput.xlsx')


#export the cells geojson data to spreadsheet. previously xlsoutput.py
def cells_geojson_to_sheet(input_file_path, output_file_path):
    with open(input_file_path) as f:
        data = json.load(f)

    # Create empty lists to store the properties and geometries
    properties = []
    geometries = []

    # Loop through each feature in the GeoJSON file
    for feature in data["features"]:
        # Get the properties and geometry of the feature
        properties.append(feature["properties"])
        geometries.append(feature["geometry"])

    # Convert the lists of properties and geometries to dataframes
    properties_df = pd.DataFrame(properties)
    geometries_df = pd.DataFrame(geometries)

    df = pd.concat([properties_df, geometries_df], axis=1)

    # Rename the columns "state" to "Kingdom" and "province" to "County"
    df.rename(columns={'state': 'Kingdom', 'province': 'County', 'religion': 'Religion', 'culture': 'Culture'},
              inplace=True)

    def hex_to_rgb(hex_color):
        if not hex_color.startswith("#") or len(hex_color) != 7:
            return None
        return tuple(int(hex_color[i:i + 2], 16) for i in (1, 3, 5))

    df["color"] = df["color"].astype(str)
    df[["red", "green", "blue"]] = df["color"].apply(hex_to_rgb).apply(pd.Series)
    df = df.drop("color", axis=1)

    names = ['Abingdon', 'Albrighton', 'Alcester', 'Almondbury', 'Altrincham', 'Amersham', 'Andover', 'Appleby',
             'Ashboume', 'Atherstone', 'Aveton', 'Axbridge', 'Aylesbury', 'Baldock', 'Bamburgh', 'Barton',
             'Basingstoke', 'Berden', 'Bere', 'Berkeley', 'Berwick', 'Betley', 'Bideford', 'Bingley', 'Birmingham',
             'Blandford', 'Blechingley', 'Bodmin', 'Bolton', 'Bootham', 'Boroughbridge', 'Boscastle', 'Bossinney',
             'Bramber', 'Brampton', 'Brasted', 'Bretford', 'Bridgetown', 'Bridlington', 'Bromyard', 'Bruton',
             'Buckingham', 'Bungay', 'Burton', 'Calne', 'Cambridge', 'Canterbury', 'Carlisle', 'Castleton', 'Caus',
             'Charmouth', 'Chawleigh', 'Chichester', 'Chillington', 'Chinnor', 'Chipping', 'Chisbury', 'Cleobury',
             'Clifford', 'Clifton', 'Clitheroe', 'Cockermouth', 'Coleshill', 'Combe', 'Congleton', 'Crafthole',
             'Crediton', 'Cuddenbeck', 'Dalton', 'Darlington', 'Dodbrooke', 'Drax', 'Dudley', 'Dunstable', 'Dunster',
             'Dunwich', 'Durham', 'Dymock', 'Exeter', 'Exning', 'Faringdon', 'Felton', 'Fenny', 'Finedon', 'Flookburgh',
             'Fowey', 'Frampton', 'Gateshead', 'Gatton', 'Godmanchester', 'Grampound', 'Grantham', 'Guildford',
             'Halesowen', 'Halton', 'Harbottle', 'Harlow', 'Hatfield', 'Hatherleigh', 'Haydon', 'Helston', 'Henley',
             'Hertford', 'Heytesbury', 'Hinckley', 'Hitchin', 'Holme', 'Hornby', 'Horsham', 'Kendal', 'Kenilworth',
             'Kilkhampton', 'Kineton', 'Kington', 'Kinver', 'Kirby', 'Knaresborough', 'Knutsford', 'Launceston',
             'Leighton', 'Lewes', 'Linton', 'Louth', 'Luton', 'Lyme', 'Lympstone', 'Macclesfield', 'Madeley',
             'Malborough', 'Maldon', 'Manchester', 'Manningtree', 'Marazion', 'Marlborough', 'Marshfield', 'Mere',
             'Merryfield', 'Middlewich', 'Midhurst', 'Milborne', 'Mitford', 'Modbury', 'Montacute', 'Mousehole',
             'Newbiggin', 'Newborough', 'Newbury', 'Newenden', 'Newent', 'Norham', 'Northleach', 'Noss', 'Oakham',
             'Olney', 'Orford', 'Ormskirk', 'Oswestry', 'Padstow', 'Paignton', 'Penkneth', 'Penrith', 'Penzance',
             'Pershore', 'Petersfield', 'Pevensey', 'Pickering', 'Pilton', 'Pontefract', 'Portsmouth', 'Preston',
             'Quatford', 'Reading', 'Redcliff', 'Retford', 'Rockingham', 'Romney', 'Rothbury', 'Rothwell', 'Salisbury',
             'Saltash', 'Seaford', 'Seasalter', 'Sherston', 'Shifnal', 'Shoreham', 'Sidmouth', 'Skipsea', 'Skipton',
             'Solihull', 'Somerton', 'Southam', 'Southwark', 'Standon', 'Stansted', 'Stapleton', 'Stottesdon',
             'Sudbury', 'Swavesey', 'Tamerton', 'Tarporley', 'Tetbury', 'Thatcham', 'Thaxted', 'Thetford', 'Thornbury',
             'Tintagel', 'Tiverton', 'Torksey', 'Totnes', 'Towcester', 'Tregoney', 'Trematon', 'Tutbury', 'Uxbridge',
             'Wallingford', 'Wareham', 'Warenmouth', 'Wargrave', 'Warton', 'Watchet', 'Watford', 'Wendover', 'Westbury',
             'Westcheap', 'Weymouth', 'Whitford', 'Wickwar', 'Wigan', 'Wigmore', 'Winchelsea', 'Winkleigh', 'Wiscombe',
             'Witham', 'Witheridge', 'Wiveliscombe', 'Woodbury', 'Yeovil'
             ]
    df['Barony'] = [random.choice(names) + random.choice(names) for i in range(len(df))]

    # Save the dataframe to an XLS spreadsheet
    df.to_excel(output_file_path, index=False)


#Takes names from the output json data and combined with cells data to generate the provinceDef.xlsx
def nameCorrector(cells_file_path, combined_file_path, updated_file_path):
    # Load the first Excel file into a pandas DataFrame
    df1 = pd.read_excel(cells_file_path)

    # Load the second Excel file into a pandas DataFrame
    df2 = pd.read_excel(combined_file_path)

    df3 = pd.read_excel(combined_file_path, sheet_name=1)
    df4 = pd.read_excel(combined_file_path, sheet_name=2)
    df5 = pd.read_excel(combined_file_path, sheet_name=3)

    # Create a dictionary from the second DataFrame that maps numbers to names
    mapping = dict(zip(df2['i'], df2['name']))
    provmapping = dict(zip(df3['i'], df3['name']))
    culturemapping = dict(zip(df4['i'], df4['name']))
    religionmapping = dict(zip(df5['i'], df5['name']))

    # Replace the numbers in the first DataFrame with the associated names
    df1['Kingdom'] = df1['Kingdom'].map(mapping)
    df1['County'] = df1['County'].map(provmapping)
    df1['Religion'] = df1['Religion'].map(religionmapping)
    df1['Culture'] = df1['Culture'].map(culturemapping)

    # Save the updated first DataFrame to a new Excel file
    df1.to_excel(updated_file_path, index=False)


#Rearranges the nameCorrector data into format needed by the Map Filler tool
def provinceDefCells(file_path, output_path):
    # Load the existing Excel file
    wb = openpyxl.load_workbook(file_path)

    # Create a new sheet
    ws = wb.create_sheet("Copied Data")

    # Get the data from the original sheet
    source_sheet = wb["Sheet1"]

    # Copy data from columns
    for i in range(1, source_sheet.max_row + 1):
        cell_value = source_sheet.cell(row=i, column=1).value
        if isinstance(cell_value, int):
            ws.cell(row=i, column=2, value=cell_value + 1)
        else:
            ws.cell(row=i, column=2, value=cell_value)
        ws.cell(row=i, column=3, value=source_sheet.cell(row=i, column=13).value)
        ws.cell(row=i, column=4, value=source_sheet.cell(row=i, column=14).value)
        ws.cell(row=i, column=5, value=source_sheet.cell(row=i, column=15).value)
        ws.cell(row=i, column=6, value=source_sheet.cell(row=i, column=16).value)
        ws.cell(row=i, column=9, value=source_sheet.cell(row=i, column=6).value)
        ws.cell(row=i, column=10, value=source_sheet.cell(row=i, column=6).value)
        ws.cell(row=i, column=11, value=source_sheet.cell(row=i, column=7).value)
        ws.cell(row=i, column=12, value=source_sheet.cell(row=i, column=7).value)

        #For Religion Name
        value = source_sheet.cell(row=i, column=9).value
        if value is not None:
            if value.lower() in ["no religion", "no_religion"]:
                value = ""  # make the cell blank
            else:
                value = re.sub(r'\W+', '', value).lower()  # remove non-alphanumeric characters and spaces
            ws.cell(row=i, column=14, value=value)

    # Save only the "Copied Data" sheet in the output file
    wb_copied_data = openpyxl.Workbook()
    ws_copied_data = wb_copied_data.active
    ws_copied_data.title = "Copied Data"
    for row in ws.iter_rows(values_only=True):
        ws_copied_data.append(row)
    wb_copied_data.save(output_path)



import os
def terrainGenIdtoName(cells_path, biomes_path):
    # Load biomes file
    biomes_df = pd.read_excel(biomes_path, header=None)

    # Create a dictionary mapping biome IDs to biome names
    biome_dict = dict(zip(biomes_df.iloc[:, 1], biomes_df.iloc[:, 0]))

    # Load cellsData file
    cells_df = pd.read_csv(cells_path)

    # Replace biome IDs with biome names using the dictionary
    cells_df['biome'] = cells_df['biome'].map(biome_dict)

    # Save the updated cellsData file to the same directory
    output_path = cells_path
    cells_df.to_csv(output_path, index=False)




def terrainGen(cellsData,provinceTerraintxt):
    import pandas as pd
    import random

    # read in the CSV file
    df = pd.read_csv(cellsData)

    # define the biome-terrain mapping with weights
    biome_terrain_map = {
        'Marine': {
            'plains': 1.0
        },
        'Hot desert': {
            'desert': 0.6,
            'desert_mountains': 0.3,
            'drylands': 0.1
        },
        'Cold desert': {
            'desert': 0.6,
            'desert_mountains': 0.3,
            'tundra': 0.1
        },
        'Savanna': {
            'plains': 0.6,
            'hills': 0.3,
            'steppe': 0.1
        },
        'Grassland': {
            'steppe': 0.9,
            'hills': 0.1
        },
        'Tropical seasonal forest': {
            'jungle': 0.4,
            'plains': 0.4,
            'hills': 0.2
        },
        'Temperate deciduous forest': {
            'forest': 0.6,
            'hills': 0.4
        },
        'Tropical rainforest': {
            'jungle': 1.0
        },
        'Temperate rainforest': {
            'forest': 0.7,
            'hills': 0.3
        },
        'Taiga': {
            'taiga': 0.8,
            'hills': 0.2
        },
        'Tundra': {
            'tundra': 0.8,
            'hills': 0.2
        },
        'Glacier': {
            'mountains': 1.0
        },
        'Wetland': {
            'wetlands': 1.0,
        }
    }

    # create a list to hold the assigned terrain for each cell
    terrain_list = []

    # loop through each row in the DataFrame and assign a terrain based on the biome
    for i, row in df.iterrows():
        biome = row['biome']
        terrain_weights = biome_terrain_map[biome]

        # check for population and override terrain for Temperate deciduous forest biomes
        if biome == 'Temperate deciduous forest' and row['population'] > 35000:
            assigned_terrain = 'farmlands'
        elif biome == 'Hot Desert' and row['population'] > 35000:
            assigned_terrain = 'floodplains'
        else:
            assigned_terrain = random.choices(list(terrain_weights.keys()), weights=list(terrain_weights.values()))[0]

        terrain_list.append(f"{row['i']}={assigned_terrain}")

    # write the assigned terrain for each cell to a text file
    with open(provinceTerraintxt, 'w') as f:
        f.write('\n'.join(terrain_list))




#remove_emoji_from_json("emoji.json", "noemoji.json")

#colorRandom("input.geojson","output.geojson")

#json_to_sheet("noemoji.json","combined_data.xlsx")

#cells_geojson_to_sheet("output.geojson","cellsData.xlsx")

#nameCorrector('cellsData.xlsx', 'combined_data.xlsx', 'updated_file.xlsx')

#provinceDefCells("updated_file.xlsx", "provinceDef.xlsx")

