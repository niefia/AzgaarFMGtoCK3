import os
import openpyxl
import random
import re



def generate_start_day():
    BM = 1066
    year = BM
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    return f"{year}.{month}.{day}"

bm_output = generate_start_day()

def bm_groups(output_dir):

    subfolder_path = os.path.join(output_dir, 'common', 'bookmarks', 'groups')
    if not os.path.exists(subfolder_path):
        os.makedirs(subfolder_path)
    filename = os.path.join(subfolder_path, '00_bookmark_group.txt')
    print(filename)

    with open(filename, 'w') as f:
        f.write(f"bm_group_{bm_output} = {{\n")
        f.write(f"\tdefault_start_date = {bm_output}\n}}")


def bm_generator(output_dir, num_chars):
    # Load the characters.xlsx file
    wb = openpyxl.load_workbook(os.path.join(output_dir, 'characters.xlsx'))
    sheet = wb.active

    # Select a random row (excluding the header row) for each character
    max_row = sheet.max_row - 1
    char_data = []
    for i in range(num_chars):
        rand_row = random.randint(2, max_row + 1)

        # Read the values for the selected row
        id_num = sheet.cell(row=rand_row, column=1).value
        name = sheet.cell(row=rand_row, column=2).value
        bdate = sheet.cell(row=rand_row, column=3).value
        culture = sheet.cell(row=rand_row, column=4).value
        religion = sheet.cell(row=rand_row, column=5).value
        religion = re.sub(r'\W+', '', religion)  # Remove non-alphanumeric characters
        religion = religion.lower()

        posx = random.randint(150, 850)
        posy = random.randint(150, 850)

        char_data.append({
            'id_num': id_num,
            'name': name,
            'bdate': bdate,
            'culture': culture,
            'religion': religion,
            'posx': posx,
            'posy': posy,
        })

    # Write the bookmark data to file
    subfolder_path = os.path.join(output_dir, 'common', 'bookmarks', 'bookmarks')
    if not os.path.exists(subfolder_path):
        os.makedirs(subfolder_path)
    filename = os.path.join(subfolder_path, '00_bookmarks.txt')
    with open(filename, 'w') as f:

        #Generate bookmark file opening data
        f.write(f"bm_{bm_output} = {{\n")
        f.write(f"\tstart_date = {bm_output}\n")
        f.write(f"\tis_playable = yes\n")
        f.write(f"\trecommended = yes\n")
        f.write(f"\tgroup = bm_group_{bm_output}\n\n")

        #Generate i Bookmark character from random characters in Characters.xlsx
        for i, char in enumerate(char_data):
            f.write(f"\tcharacter = {{\n")
            f.write(f'\t\tname = "bookmark_{char["name"]}"\n')
            f.write(f"\t\thistory_id = {char['id_num']}\n")
            f.write(f"\t\tbirth = {char['bdate']}\n")
            f.write(f"\t\tculture = {char['culture']}\n")
            f.write(f"\t\treligion = {char['religion']}\n")
            f.write(f"\t\ttitle = k_{char['name'].lower()}\n")
            f.write(f'\t\tdifficulty = "BOOKMARK_CHARACTER_DIFFICULTY_EASY"\n')
            f.write(f"\t\tposition = {{{char['posx']} {char['posy']}}}\n")
            f.write(f"\t}}\n")

        f.write(f"}}\n")

#bm_groups(output_dir)
#bm_generator(output_dir, 3)