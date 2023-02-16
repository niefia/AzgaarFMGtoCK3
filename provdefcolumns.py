import openpyxl
import os
import re

# Load the existing Excel file
wb = openpyxl.load_workbook("updated_file.xlsx")

# Create a new sheet
ws = wb.create_sheet("Copied Data")

# Get the data from the original sheet
source_sheet = wb["Sheet1"]

# Copy data from columns
for i in range(1, source_sheet.max_row + 1):
    cell_value = source_sheet.cell(row=i, column=1).value
    if isinstance(cell_value, int):
        ws.cell(row=i, column=2, value=cell_value + 1)
    else:
        ws.cell(row=i, column=2, value=cell_value)
    ws.cell(row=i, column=3, value=source_sheet.cell(row=i, column=13).value)
    ws.cell(row=i, column=4, value=source_sheet.cell(row=i, column=14).value)
    ws.cell(row=i, column=5, value=source_sheet.cell(row=i, column=15).value)
    ws.cell(row=i, column=6, value=source_sheet.cell(row=i, column=16).value)
    ws.cell(row=i, column=9, value=source_sheet.cell(row=i, column=6).value)
    ws.cell(row=i, column=10, value=source_sheet.cell(row=i, column=6).value)
    ws.cell(row=i, column=11, value=source_sheet.cell(row=i, column=7).value)
    ws.cell(row=i, column=12, value=source_sheet.cell(row=i, column=7).value)

    #For Religion Name
    value = source_sheet.cell(row=i, column=9).value
    if value is not None:
        if value.lower() in ["no religion", "no_religion"]:
            value = ""  # make the cell blank
        else:
            value = re.sub(r'\W+', '', value)  # remove non-alphanumeric characters and spaces
        ws.cell(row=i, column=14, value=value)




# Save only the "Copied Data" sheet in the "provinceDef.xlsx" file
wb_copied_data = openpyxl.Workbook()
ws_copied_data = wb_copied_data.active
ws_copied_data.title = "Copied Data"

for row in ws.iter_rows(values_only=True):
    ws_copied_data.append(row)

wb_copied_data.save("provinceDef.xlsx")
