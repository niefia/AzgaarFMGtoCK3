import pandas as pd
from openpyxl import Workbook
import openpyxl
import re

#Rearrange and extract BFS Data into new provinceDef file (Cell ID is not used in game but used later to get parent cell data on province,state, religion culture etc)
def extractBFS():
    # Load the data from the BFSoutput.xlsx file into a pandas DataFrame
    df = pd.read_excel("BFSoutput.xlsx")

    # Filter the DataFrame to only include rows where distance = 0
    df = df[df["distance"] == 0]

    # Convert the color column from hex to rgb and split into separate columns for r, g, and b
    df[["R", "G", "B"]] = df["color"].apply(lambda x: pd.Series(tuple(int(x[i:i+2], 16) for i in (0, 2, 4)) if len(x) == 6 else (pd.NA, pd.NA, pd.NA)))

    # Create a new workbook and worksheet for writing the filtered data to a new Excel file
    wb = Workbook()
    ws = wb.active

    # Write the header row to the worksheet
    header = ["", "", "R", "G", "B", "Nearest Town", "Cell ID"]
    ws.append(header)

    # Write the filtered data to the worksheet
    for row in df.itertuples(index=False):
        ws.append(["", "", row.R, row.G, row.B, row.nearest_town,row.id])

    # Save the new Excel file
    wb.save("provinceDef.xlsx")

#Copy over Town Id's which will be used as Barony id
def BaronyId():
    import pandas as pd

    # Load the "combined_data.xlsx" file and select the "burgs" sheet
    combined_data = pd.read_excel("combined_data.xlsx", sheet_name="burgs")

    # Load the "provinceDef.xlsx" file
    province_def = pd.read_excel("provinceDef.xlsx")

    # Copy the contents of column 0 from the "combined_data" DataFrame to column 1 of the "province_def" DataFrame
    province_def[province_def.columns[1]] = combined_data.iloc[:, 0]

    # Save the updated "province_def" DataFrame to "provinceDef.xlsx"
    province_def.to_excel("provinceDef.xlsx", index=False)

#Uses the Cell ID obtained earlier to assign Province data
def ProvData():
    # Load the two Excel files into Pandas dataframes
    updated_file = pd.read_excel('updated_file.xlsx')
    province_def = pd.read_excel('provinceDef.xlsx')

    # Merge the two dataframes based on the 'Cell ID' and 'id' columns
    merged = pd.merge(province_def, updated_file, left_on='Cell ID', right_on='id')

    # Replace the 'Cell ID' column with the 'type' column
    merged['Cell ID'] = merged['type']

    # Define the columns to transfer
    columns_to_transfer = ['Cell ID', 'population', 'Kingdom', 'County', 'Culture', 'Religion']

    # Apply re.sub() function to the 'Religion' column
    merged['Religion'] = merged['Religion'].apply(lambda x: re.sub(r'\W+', '', x) if isinstance(x, str) else x)

    # Transfer the columns from updated_file to province_def for the matching rows
    province_def.loc[merged.index, columns_to_transfer] = merged[columns_to_transfer]

    # Save the modified province_def dataframe to a new Excel file
    province_def.to_excel('provinceDef.xlsx', index=False)

#Reorder Columns to provinceDef order
def cOrder():
    import openpyxl

    # Open the Excel file
    wb = openpyxl.load_workbook('provinceDef.xlsx')

    # Select the active worksheet
    ws = wb.active

    # Get the column index of the "Religion", "Culture", "County", and "Kingdom" columns
    religion_col_idx = 0
    culture_col_idx = 0
    county_col_idx = 0
    kingdom_col_idx = 0
    for cell in ws[1]:
        if cell.value == "Religion":
            religion_col_idx = cell.column
        elif cell.value == "Culture":
            culture_col_idx = cell.column
        elif cell.value == "County":
            county_col_idx = cell.column
        elif cell.value == "Kingdom":
            kingdom_col_idx = cell.column

    # Move the "Religion", "Culture", "County", and "Kingdom" columns to columns M, N, O, and P, respectively
    if religion_col_idx > 0:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(religion_col_idx)}1:{openpyxl.utils.get_column_letter(religion_col_idx)}{ws.max_row}",
            cols=2)
    if culture_col_idx > 0:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(culture_col_idx)}1:{openpyxl.utils.get_column_letter(culture_col_idx)}{ws.max_row}",
            cols=2)
    if county_col_idx > 0:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(county_col_idx)}1:{openpyxl.utils.get_column_letter(county_col_idx)}{ws.max_row}",
            cols=2)
    if kingdom_col_idx > 0:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(kingdom_col_idx)}1:{openpyxl.utils.get_column_letter(kingdom_col_idx)}{ws.max_row}",
            cols=1)

    # Save the updated Excel file
    wb.save('provinceDef.xlsx')

#Copy Kingdoms to Empire, Counties to Duchies & Delete Cell ID and population
def cOrder2():

        # Load the workbook
    workbook = openpyxl.load_workbook('provinceDef.xlsx')

    # Select the active worksheet
    worksheet = workbook.active

    # Loop through each row
    for i in range(1, worksheet.max_row + 1):
        # Copy column 10 to column 9
        worksheet.cell(row=i, column=9).value = worksheet.cell(row=i, column=10).value

        # Delete the values in columns 7
        worksheet.cell(row=i, column=7).value = None

        # Copy column 12 to column 11
        worksheet.cell(row=i, column=11).value = worksheet.cell(row=i, column=12).value

    # Save the changes to the workbook
    workbook.save('provinceDef.xlsx')



extractBFS()
BaronyId()
ProvData()
cOrder()
#cOrder2()
