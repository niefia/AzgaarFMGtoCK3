import ast
import collections
import os
from queue import Queue
import pandas as pd
import openpyxl
from openpyxl import Workbook, utils
from openpyxl.styles import PatternFill
import random
import hextorgb
from PIL import Image, ImageDraw
import re
import xlwt
import time


def bfs_distance(combined_data_file, cells_data_file, output_file, useCounties):
    # Read data from Excel files
    combined_data = pd.read_excel(combined_data_file, sheet_name="burgs")
    cells_data = pd.read_excel(cells_data_file, usecols=["id", "neighbors", "coordinates", "type", "County"])
    cells_data['neighbors'] = cells_data['neighbors'].apply(lambda x: [int(n) for n in x.strip('[]').split(',')])
    # Create a dictionary to map cell IDs to town names
    id_to_town = dict(zip(combined_data["cell"], combined_data["name"]))
    # Define starting points as cells containing towns
    start = combined_data["cell"].tolist()
    # Define the frontier, cost_so_far, and started_at dictionaries
    frontier = Queue()
    cost_so_far = dict()
    started_at = dict()
    coordinates = dict()
    cellType = dict()
    neighbors = dict()
    impassable = dict()
    # Starting points get distance 0 and will point to themselves
    for location in start:
        frontier.put(location)
        cost_so_far[location] = 0
        started_at[location] = location
        cellType[location] = cells_data.loc[cells_data["id"] == location, "type"].tolist()[0]
        coordinates[location] = cells_data.loc[cells_data["id"] == location, "coordinates"].tolist()[0]
        neighbors[location] = cells_data.loc[cells_data["id"] == location, "neighbors"].tolist()[0]
        impassable[location] = False
    # Expand outwards from existing points
    while not frontier.empty():
        current = frontier.get()        
        # Idea: Make impassable when too large
        if cost_so_far[current] >= 10:
            impassable[current] = True
        else:
            impassable[current] = False
        current_cell_County = cells_data.loc[cells_data["id"] == current, "County"].tolist()[0]
        current_cell_Type = cells_data.loc[cells_data["id"] == current, "type"].tolist()[0]
        # cellType[location] = current_cell_Type
        # if current in cells_data["id"].tolist() and (current_cell_Type != "ocean" and current_cell_Type != "lake"):
        if current in cells_data["id"].tolist():
            for next_cell in neighbors[current]:
                if (next_cell not in cost_so_far):
                    next_cell_Type = cells_data.loc[cells_data["id"] == next_cell, "type"].tolist()[0]
                    next_cell_County = cells_data.loc[cells_data["id"] == next_cell, "County"].tolist()[0]
                    # next_cell_Coordinates = cells_data.loc[cells_data["id"] == next_cell, "coordinates"].tolist()[0]                                         
                    # # Comment snibbo: write neighbors
                    # if (next_cell_County != current_cell_County) and (next_cell not in neighbors[current]) and (next_cell_Type != "ocean") and (next_cell_Type != "lake"):
                    #     neighbors[current].append(next_cell)
                        # Room for improvement: find neighbors across the sea                    
                    if (next_cell_Type != "ocean" and next_cell_Type != "lake") and (not useCounties or next_cell_County == current_cell_County):
                        started_at[next_cell] = started_at[current]
                        cost_so_far[next_cell] = cost_so_far[current] + 1
                        cellType[next_cell] = next_cell_Type
                        next_cell_Coordinates = cells_data.loc[cells_data["id"] == next_cell, "coordinates"].tolist()[0]
                        coordinates[next_cell] = next_cell_Coordinates
                        next_cell_Neighbors = cells_data.loc[cells_data["id"] == next_cell, "neighbors"].tolist()[0]                     
                        neighbors[next_cell] = next_cell_Neighbors
                        frontier.put(next_cell)

    # Create a DataFrame to store the results
    output_data = pd.DataFrame(list(cost_so_far.items()), columns=["id", "distance"])
    # Add a column to the output DataFrame for the nearest town and coordinates
    output_data["nearest_town"] = output_data["id"].apply(lambda x: id_to_town[started_at[x]])
    output_data["coordinates"] = output_data["id"].apply(lambda x: coordinates.get(x))
    output_data["cellType"] = output_data["id"].apply(lambda x: cellType.get(x))
    output_data["impassable"] = output_data["id"].apply(lambda x: impassable.get(x))    
    output_data["neighbors"] = output_data["id"].apply(lambda x: neighbors.get(x))

    # Save the output DataFrame to a new Excel file
    output_data.to_excel(output_file, index=False)
    




def colorRandomBFS(file_name):
    # Open the Excel file and select the active worksheet
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active

    # Get the values from the 'nearest_town' column
    values = [cell.value for cell in worksheet['C'][1:]]

    # Find the unique values and assign a random color to each one
    unique_values = list(set(values))
    color_dict = {}
    for value in unique_values:
        color_dict[value] = '{:06x}'.format(random.randint(0, 0xFFFFFF))

    # # Check if column 'color' exists
    # if 'color' in worksheet.columns:
    #     intColColor = openpyxl.utils.column_index_from_string('color')
    # else:
    #     # Rename the column to 'color'
    #     print("Column A does not exist.")
    #     intColColor = 9
    #     worksheet.cell(row=1, column=intColColor).value = "color"
    intColColor = 9
    worksheet.cell(row=1, column=intColColor).value = "color"

    # Add a new column called 'color' and assign each row a color based on its 'nearest_town' value
    for i, value in enumerate(values):
        cell = worksheet.cell(row=i+2, column=intColColor)    # comment snibbo: changed to 8 due to new columns
        cell.value = color_dict[value]
        cell.fill = PatternFill(start_color=color_dict[value], end_color=color_dict[value], fill_type='solid')

    # Save the updated worksheet to a new sheet called 'Output'
    workbook.save(file_name)


def provinceMapBFS(file_path, scaling_factor,output_folder):
    # read the Excel file
    df = pd.read_excel(file_path)

    # extract coordinates and colors
    coords_str = df['coordinates'].values
    colors_str = df['color'].values
    coords = []
    colors = []
    for coord_str, color_str in zip(coords_str, colors_str):
        coord_list = ast.literal_eval(coord_str)
        coord_list = [[float(x), float(y)] for x, y in coord_list[0]]
        coords.append(coord_list)
        colors.append(color_str)

    # set up image
    width, height = 8192, 4096
    img = Image.new('RGBA', (width, height), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    # draw each polygon with scaling factor applied
    for coord, color in zip(coords, colors):

        # calculate the center pixel
        center_px = (width / 2, height / 2)

        # calculate the pixel value for each coordinate, with scaling factor applied and centered at the center_px
        coord_px = [(int(center_px[0] + (x * scaling_factor)),
                     int(center_px[1] - (y * scaling_factor))) for x, y in coord]

        # draw polygon
        rgbcolor = hextorgb.bfs_hex_to_rgb(color)
        draw.polygon(coord_px, fill=rgbcolor, outline=None)

    img.show()
    # save image
    img.save(output_folder)


#Rearrange and extract BFS Data into new provinceDef file (Cell ID is not used in game but used later to get parent cell data on province,state, religion culture etc)
def extractBFS(input_file, output_file):
    # Load the data from the input file into a pandas DataFrame
    df = pd.read_excel(input_file)
    # Filter the DataFrame to only include rows where distance = 0
    df = df[df["distance"] == 0]
    # Convert the color column from hex to rgb and split into separate columns for r, g, and b
    df[["R", "G", "B"]] = df["color"].apply(lambda x: pd.Series(tuple(int(x[i:i+2], 16) for i in (0, 2, 4)) if len(x) == 6 else (pd.NA, pd.NA, pd.NA)))
    # Create a new workbook and worksheet for writing the filtered data to a new Excel file
    wb = Workbook()
    ws = wb.active
    # Write the header row to the worksheet
    header = ["", "", "R", "G", "B", "Nearest Town", "Cell ID"]
    ws.append(header)
    # Write the filtered data to the worksheet
    for row in df.itertuples(index=False):
        ws.append(["", "", row.R, row.G, row.B, row.nearest_town,row.id])
    # Save the new Excel file with the provided output file name
    wb.save(output_file)


#Copy over Town Id's which will be used as Barony id
def BaronyId(input_file, output_file):
    # Load the input file and select the "burgs" sheet
    combined_data = pd.read_excel(input_file, sheet_name="burgs")
    # Load the "provinceDef.xlsx" file
    province_def = pd.read_excel(output_file)
    # Copy the contents of column 0 from the "combined_data" DataFrame to column 1 of the "province_def" DataFrame
    province_def[province_def.columns[1]] = combined_data.iloc[:, 0]
    # Save the updated "province_def" DataFrame to the specified output file
    province_def.to_excel(output_file, index=False)

def BaronyIdBiomes(combined_data, cellsData, townbiomes):
    # Open the Excel files
    combined_data = pd.read_excel(combined_data, sheet_name='burgs')
    cells_data = pd.read_excel(cellsData)
    # Merge the two dataframes on the 'id' and 'cells' columns
    merged_data = pd.merge(combined_data, cells_data[['id', 'biome', 'population']], left_on='cell', right_on='id')
    # Add the 'biome' and 'population' columns to the 'burgs' sheet
    combined_data['biome'] = merged_data['biome']
    combined_data['population'] = merged_data['population']
    # Save the updated data to townBiomes.csv
    combined_data.to_csv(townbiomes, index=False)


def ProvData(updated_file_name, province_def_name, output_file_name):
    # Load the two Excel files into Pandas dataframes
    updated_file = pd.read_excel(updated_file_name)
    province_def = pd.read_excel(province_def_name)
    # Merge the two dataframes based on the 'Cell ID' and 'id' columns
    merged = pd.merge(province_def, updated_file, left_on='Cell ID', right_on='id')
    # Define the columns to transfer
    columns_to_transfer = ['Cell ID', 'population', 'Kingdom', 'County', 'Culture', 'Religion']
    # Create Religion column
    merged['Religion'] = merged['Religion'].apply(lambda x: x.lower() if isinstance(x, str) else x)
    # Apply re.sub() function to the 'Religion' column
    merged['Religion'] = merged['Religion'].apply(lambda x: re.sub(r'\W+', '', x) if isinstance(x, str) else x)
    # Transfer the columns from updated_file to province_def for the matching rows
    province_def.loc[merged.index, columns_to_transfer] = merged[columns_to_transfer]
    # Save the modified province_def dataframe to the specified output file
    province_def.to_excel(output_file_name, index=False)


def convert_xlsx_to_xls(excel_file_path, output_file):
    # Open the input .xlsx file
    wb = openpyxl.load_workbook(excel_file_path)
    # Create a new .xls workbook
    xls_wb = xlwt.Workbook()
    # Copy each worksheet from the .xlsx workbook to the new .xls workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        xls_sheet = xls_wb.add_sheet(sheet_name)
        # Loop over .xlsx and copy each row into .xls
        for row_index, row in enumerate(sheet.iter_rows()):
            for col_index, cell in enumerate(row):
                xls_sheet.write(row_index, col_index, cell.value)
    # Save the new .xls file
    xls_wb.save(output_file)


# This function writes new Baronies to "Burgs" when a Barony is too large. This slows down the converter heavily, because it has to re-iterate the distance calculation multiple times. Maybe this should be an optional parameter in the GUI.
def bfs_distance_to_new_baronies(combined_data_file, cells_data_file, maxSizeBarony, useCounties):
    print("bfs_distance_to_new_baronies - started: ", time.ctime())
    # Read data from Excel files
    combined_data = pd.read_excel(combined_data_file, sheet_name="burgs")
    cells_data = pd.read_excel(cells_data_file, usecols=["id", "neighbors", "type", "County", "Barony"])
    # Split neighbors from cells  
    cells_data['neighbors'] = cells_data['neighbors'].apply(lambda x: [int(n) for n in x.strip('[]').split(',')])
    # Create a dictionary to map cell IDs to town names
    id_to_town = dict(zip(combined_data["cell"], combined_data["name"]))
    # Define starting points as cells containing towns
    start = combined_data["cell"].tolist()
    # Define the frontier, cost_so_far, and started_at dictionaries
    frontier = Queue()
    cost_so_far = dict()
    started_at = dict()
    neighbors = dict()    
    # Starting points get distance 0 and will point to themselves
    for location in start:
        frontier.put(location)
        cost_so_far[location] = 0
        started_at[location] = location
        neighbors[location] = []
    # Expand outwards from existing points    
    while not frontier.empty():
        # Get new Frontier from list
        current = frontier.get()
        # Get cell information 
        current_cell_County = cells_data.loc[cells_data["id"] == current, "County"].tolist()[0]
        current_cell_neighbors = cells_data.loc[cells_data["id"] == current, "neighbors"].tolist()[0]

        # Check if a Barony has reached its max size
        # This works, but takes very long...
        if cost_so_far[current] > maxSizeBarony:
            # Get last Id from current Burg-List
            lastId = combined_data.iloc[-1]["i"]
            # Create a new name >> TODO: This should be improved
            i = 1
            new_name = id_to_town[started_at[current]] + "-" + str(i)
            while new_name in id_to_town.values():
                i += 1
                new_name = id_to_town[started_at[current]] + "-" + str(i)
            # Write new Burg to Burg-List
            list = [lastId + 1, current, new_name, "random (to be generated)"]
            combined_data.loc[len(combined_data)] = list
            # Save the output DataFrame to a new Excel file
            with pd.ExcelWriter(combined_data_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:  
                combined_data.to_excel(writer, sheet_name='burgs', index=False)
            # Restart the process 
            print("new barony added (" + new_name + ") - restart calculating for distances", lastId + 1)
            bfs_distance_to_new_baronies(combined_data_file, cells_data_file, maxSizeBarony, useCounties)
            return None
            
        # Make sure that the new current cell is in the Cells-List 
        if current in cells_data["id"].tolist():
            for next_cell in current_cell_neighbors:
                if (next_cell not in cost_so_far):
                    next_cell_Type = cells_data.loc[cells_data["id"] == next_cell, "type"].tolist()[0]
                    next_cell_County = cells_data.loc[cells_data["id"] == next_cell, "County"].tolist()[0]
                    # Comment snibbo: write neighbors
                    if (next_cell_County != current_cell_County) and (next_cell not in neighbors[current]) and (next_cell_Type != "ocean") and (next_cell_Type != "lake"):
                        neighbors[current].append(next_cell)
                        # Room for improvement: find neighbors across the sea                                        
                    if (next_cell_Type != "ocean" and next_cell_Type != "lake") and (not useCounties or next_cell_County == current_cell_County):
                        cost_so_far[next_cell] = cost_so_far[current] + 1
                        started_at[next_cell] = started_at[current]
                        neighbors[next_cell] = neighbors[current]
                        frontier.put(next_cell)
    # There are many cases where a county from Azgaar has no city => in this case we have to create a new one and run the function again
    for nb in neighbors:
        nbList = neighbors[nb]
        for id in nbList:
            if id not in neighbors:
                # Get last Id from current Burg-List
                lastId = combined_data.iloc[-1]["i"]
                # Create a new name >> TODO: This should be improved
                i = 0
                new_name = cells_data.loc[cells_data["id"] == id, "Barony"].tolist()[0]
                while new_name in id_to_town.values():
                    i += 1
                    new_name = cells_data.loc[cells_data["id"] == id, "Barony"].tolist()[0] + "-" + str(i)
                # Write new Burg to Burg-List
                list = [lastId + 1, id, new_name, "random (to be generated)"]
                combined_data.loc[len(combined_data)] = list
                # Save the output DataFrame to a new Excel file
                with pd.ExcelWriter(combined_data_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:  
                    combined_data.to_excel(writer, sheet_name='burgs', index=False)
                # Restart the process
                print("new barony added (" + new_name + ") - restart calculating for distances", lastId + 1)
                bfs_distance_to_new_baronies(combined_data_file, cells_data_file, maxSizeBarony)
                return None

def bfs_appendneighbors(cells_data_file, bfs_distance_file, bfs_distance_file_with_neighbors):
    # maybe we should add a function here or before to loop over "realm_Ids" (see near end of this function). A good realm size is around 20 - large baronies (which can even occure with low distance e.g. 6)

    # TODO: neighbors needs to be "two sided" - with the current solution it might be possible that a neighbour will exists only after a new frontier was set - but not sure
    bfs_distance = pd.read_excel(bfs_distance_file, sheet_name="Sheet1")
    print(time.ctime(), "Data loaded into df")
    bfs_distance['neighbors'] = bfs_distance['neighbors'].apply(ast.literal_eval) #convert to list type
    towns = bfs_distance["nearest_town"].unique()
    for t in towns:
        neighborList = bfs_distance.loc[bfs_distance["nearest_town"] == t, "neighbors"].tolist()
        # 1st: Ensure each neighbor is listed only once >> reduce list
        joinedlist = []
        for nb in neighborList:
            if nb not in joinedlist:
                joinedlist += nb
        neighborList = list(set(joinedlist))
        neighborList = sorted(neighborList)
        # write neighbors to item
        bfs_distance.loc[bfs_distance['nearest_town'] == t, 'neighbors'] = str(neighborList)

        # 2nd: write Id-Range of Barony
        idList = bfs_distance.loc[bfs_distance["nearest_town"] == t, "id"].tolist()
        bfs_distance.loc[bfs_distance['nearest_town'] == t, 'realm_Ids'] = str(idList)
    
    # TODO: Here we could / should check if a Barony is maybe to small (what is min number of cells a barony requires? is cell size changing is azgaar or setting changes?)
    # # # # Save the output DataFrame to a new Excel file
    bfs_distance.to_excel(bfs_distance_file_with_neighbors, index=False)


def bfs_putNeighborsToTowns(cells_data_file, bfs_distance_file_with_neighbors, provinceDef_file, provinceDef_file_transformed):
    # This function adds the individual towns = Baronies >> reducing all cell neighbors to only neighboring towns
    dfCellsData = pd.read_excel(cells_data_file, sheet_name="Sheet1", usecols=["id", "type"])    
    dfDistance = pd.read_excel(bfs_distance_file_with_neighbors, sheet_name="Sheet1")
    dfProvices = pd.read_excel(provinceDef_file, sheet_name="Sheet1")
    print(time.ctime(), "transforming data")
    # Finding: the function "ast.literal_eval" takes quite long...    
    dfDistance['neighbors'] = dfDistance['neighbors'].apply(ast.literal_eval) #convert to list type
    dfDistance['realm_Ids'] = dfDistance['realm_Ids'].apply(ast.literal_eval) #convert to list type    
    cellList = dfProvices['Cell ID'].tolist()
    print(time.ctime(), "transforming data 2")
    # provinceMap = dict(zip(dfProvices["id"], dfProvices["realm_Ids"]))
    # print(cellList)
    for id in dfProvices["Cell ID"].tolist():
        # print('id: ', id)
        # copy columns from distance-file
        realmList = dfDistance.loc[dfDistance["id"] == id, "realm_Ids"].tolist()[0]
        dfProvices.loc[dfProvices['Cell ID'] == id, 'realm_Ids'] = str(realmList)
        neighborList = dfDistance.loc[dfDistance["id"] == id, "neighbors"].tolist()[0]
        dfProvices.loc[dfProvices['Cell ID'] == id, 'neighbors'] = str(neighborList)

    dfProvices['realm_Ids'] = dfProvices['realm_Ids'].apply(ast.literal_eval) #convert to list type
    dfProvices['neighbors'] = dfProvices['neighbors'].apply(ast.literal_eval) #convert to list type

    print(time.ctime(), "transforming data finished")
    i = 0
    for id in dfProvices["Cell ID"].tolist():     
        i += 1
        print("search for NeighbourTowns", i)
        # if i == 5:
        #     break    # break here
        # print('start neighbour search')
        newNeighbourTowns = []
        neighborList = dfProvices.loc[dfProvices["Cell ID"] == id, "neighbors"].tolist()[0]
        realmList = dfProvices.loc[dfProvices["Cell ID"] == id, "realm_Ids"].tolist()[0]
        for nb in neighborList:
            # check if neighboar is part of realm
            if nb not in realmList:
                # check if ocean?
                cellType = dfCellsData.loc[dfCellsData["id"] == nb, "type"].tolist()[0]
                # print(nb, "not in realm")
                # print(cellType, type(cellType))
                if cellType != "ocean" and cellType != "lake": 
                    for id2 in dfProvices["Cell ID"].tolist():
                        realmList2 = []
                        if id != id2:
                            # print('id1', id)
                            # print('id2', id2)
                            realmList2 = dfProvices.loc[dfProvices["Cell ID"] == id2, "realm_Ids"].tolist()[0]
                            # print('nb', nb)
                            # print(realmList)
                            if nb in realmList2:
                                if id2 not in newNeighbourTowns:
                                    # print(id, " found another on: ", nb, " on id ", id2)
                                    newNeighbourTowns.append(id2)
                                    # newNeighbourTowns += nb
                            newNeighbourTowns = list(set(newNeighbourTowns))
                                # print("say what")
                                # print(nb)
                                # return
        dfProvices.loc[dfProvices['Cell ID'] == id, 'newNeighbourTowns'] = str(newNeighbourTowns)
        dfProvices.loc[dfProvices['Cell ID'] == id, 'countNeighbourTowns'] = len(newNeighbourTowns)

        # id = dfProvices.loc[dfProvices.loc["Cell ID"] == nb, "Nearest Town"]
        #     print(id)
        # return
    print(
        dfProvices
    )


    # Here, we could check if a barony 

    # # # # Save the output DataFrame to a new Excel file
    dfProvices.to_excel(provinceDef_file_transformed, index=False)


def bfs_calcType(biomes_file, cells_data_file, provinceDef_file):
    # get Biomes Definition from Azgaar    
    dfBiomes = pd.read_excel(biomes_file, sheet_name="Biomes", header=None, index_col=2)
    # Assign a name to the description column
    dfBiomes.columns = ['name', 'id']
    # print(dfBiomes)

    # Assign Cell Type / Biomes 
    print("it would be good if we calculate the baronie from the number of biomes entries - therefore we would have to add the biomes in 'bfs_distance' first ")
    # currently we only use type of capital
    dfCellsData = pd.read_excel(cells_data_file, sheet_name="Sheet1", usecols=["id", "biome"])
    dfProvices = pd.read_excel(provinceDef_file, sheet_name="Sheet1")    

    for id in dfProvices["Cell ID"].tolist():
        biomes =  dfCellsData.loc[dfCellsData['id'] == id, 'biome'].tolist()[0]        
        dfProvices.loc[dfProvices['Cell ID'] == id, 'biomes'] = biomes
        bio = dfBiomes.loc[dfBiomes['id'] == biomes, 'name'].tolist()[0]
        dfProvices.loc[dfProvices['Cell ID'] == id, 'biomesName'] = bio

    dfProvices.to_excel(provinceDef_file, index=False)


def cOrder(file_path):
    # Open the Excel file
    wb = openpyxl.load_workbook(file_path)
    # Select the active worksheet
    ws = wb.active  # Comment snibbo: you should define the sheet instead of using the active sheet in case Excel was opened before
    # Get the column index of the "Religion", "Culture", "County", and "Kingdom" columns
    religion_col_idx = 0
    culture_col_idx = 0
    county_col_idx = 0
    kingdom_col_idx = 0
    for cell in ws[1]:
        if cell.value == "Religion":
            religion_col_idx = cell.column
        elif cell.value == "Culture":
            culture_col_idx = cell.column
        elif cell.value == "County":
            county_col_idx = cell.column
        elif cell.value == "Kingdom":
            kingdom_col_idx = cell.column
    # Move the "Religion", "Culture", "County", and "Kingdom" columns to columns M, N, O, and P, respectively
    if religion_col_idx > 0:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(religion_col_idx)}1:{openpyxl.utils.get_column_letter(religion_col_idx)}{ws.max_row}",
            cols=2)
    if culture_col_idx > 0:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(culture_col_idx)}1:{openpyxl.utils.get_column_letter(culture_col_idx)}{ws.max_row}",
            cols=2)
    if county_col_idx > 0:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(county_col_idx)}1:{openpyxl.utils.get_column_letter(county_col_idx)}{ws.max_row}",
            cols=2)
    if kingdom_col_idx > 0:
        ws.move_range(
            f"{openpyxl.utils.get_column_letter(kingdom_col_idx)}1:{openpyxl.utils.get_column_letter(kingdom_col_idx)}{ws.max_row}",
            cols=1)
    ws["A1"] = "0"
    ws["B1"] = "ExcelIdx"
    ws["G1"] = "Cell ID"
    ws["O1"] = "Empire"
    ws["Q1"] = "Duchy"
    # Save the updated Excel file
    wb.save(file_path)


def assignEmpireAndDuchyId(provinceDef_file_transformed, duchyDef_file):
    dfBaronies = pd.read_excel(provinceDef_file_transformed, sheet_name="Sheet1")
    dfBaronies['newNeighbourTowns'] = dfBaronies['newNeighbourTowns'].apply(ast.literal_eval) #convert to list type
    # print(dfBaronies)

    # Reduce Counties with only 1 or 2 Baronies
    counties = dfBaronies["County"].unique()
    countMapCounties = dict()

    for c in counties:
        baronyList = dfBaronies.loc[dfBaronies['County'] == c, 'Cell ID'].tolist()
        countMapCounties[c] = len(baronyList)
    # print("countMapCounties")
    # print(countMapCounties)
    countMapCounties = dict(sorted(countMapCounties.items(), key=lambda x: x[1]))    
    
    unvalidEntries = dfBaronies.loc[dfBaronies['County'].isna() == True, 'Cell ID'].tolist()
    for u in unvalidEntries: 
        moved = False
        print(u)                
        newNeighbourTowns = dfBaronies.loc[dfBaronies['Cell ID'] == u, 'newNeighbourTowns'].tolist()[0]
        counterAllowedSizes = 3
        while moved == False and counterAllowedSizes != 10:
            for nb in newNeighbourTowns:
                # print(nb)
                countyName = dfBaronies.loc[dfBaronies['Cell ID'] == nb, 'County'].tolist()[0]
                kingdomName = dfBaronies.loc[dfBaronies['Cell ID'] == nb, 'Kingdom'].tolist()[0]
                # print(countyName)
                currentCount = countMapCounties[countyName]
                # print(currentCount)
                if currentCount < counterAllowedSizes:
                    dfBaronies.loc[dfBaronies['Cell ID'] == u, 'County'] = countyName
                    dfBaronies.loc[dfBaronies['Cell ID'] == u, 'Kingdom'] = kingdomName
                    countMapCounties[countyName] += 1
                    moved = True
                    break
            counterAllowedSizes += 1
        if moved == False:
            print("Is this an island?")
            print(u, newNeighbourTowns)    # id
            # In case there is no neighbour - this will be a one size barony > now it will be deleted...
            # it would be better to create a new name, but how to find the closest neighbour duchy? / kingdom?
            # dfBaronies.loc[dfBaronies['Cell ID'] == u, 'County'] = countyName
            # dfBaronies.loc[dfBaronies['Cell ID'] == u, 'Kingdom'] = kingdomName
            # countMapCounties[countyName] += 1
    # Update unvalids
    unvalidEntries = dfBaronies.loc[dfBaronies['County'].isna() == True, 'Cell ID'].tolist()
    print("still having in", unvalidEntries)
    if len(unvalidEntries) > 0:
        uName = dfBaronies.loc[dfBaronies['Cell ID'] == unvalidEntries[0], 'County'].tolist()[0]
        print(uName)        
        del countMapCounties[uName]
        
    nameStorage = []
    # print(countMapCounties)
    for c in countMapCounties:
        moved = False
        cName = c
        cCount = countMapCounties[c]
        if cCount == 1:
            # store old name
            if c not in nameStorage:
                nameStorage.append(c)            
            # remove old name and put neighbors name
            moved = False
            cellId = dfBaronies.loc[dfBaronies['County'] == c, 'Cell ID'].tolist()[0] 
            newNeighbourTowns = dfBaronies.loc[dfBaronies['Cell ID'] == cellId, 'newNeighbourTowns'].tolist()[0]

            counterAllowedSizes = 3
            while moved == False and counterAllowedSizes != 10:
                for nb in newNeighbourTowns:
                    # print(nb)
                    countyName = dfBaronies.loc[dfBaronies['Cell ID'] == nb, 'County'].tolist()[0]
                    kingdomName = dfBaronies.loc[dfBaronies['Cell ID'] == nb, 'Kingdom'].tolist()[0]
                    # print(countyName)
                    currentCount = countMapCounties[countyName]
                    # print(currentCount)
                    if currentCount < counterAllowedSizes:
                        dfBaronies.loc[dfBaronies['Cell ID'] == cellId, 'County'] = countyName
                        dfBaronies.loc[dfBaronies['Cell ID'] == cellId, 'Kingdom'] = kingdomName
                        countMapCounties[countyName] += 1
                        # del countMapCounties[c]
                        moved = True
                        break            
                counterAllowedSizes += 1
            if moved == False:
                print("Is this an island?")
                print(cellId, newNeighbourTowns)

            # assignToNeighbor(4, newNeighbourTowns)
            # assignToNeighbor(5, newNeighbourTowns)

            # for nb in newNeighbourTowns:
            #     # print(nb)
            #     countyName = dfBaronies.loc[dfBaronies['Cell ID'] == nb, 'County'].tolist()[0]
            #     kingdomName = dfBaronies.loc[dfBaronies['Cell ID'] == nb, 'Kingdom'].tolist()[0]
            #     # print(countyName)
            #     currentCount = countMapCounties[countyName]
            #     # print(currentCount)
            #     if currentCount < 3:
            #         dfBaronies.loc[dfBaronies['Cell ID'] == cellId, 'County'] = countyName
            #         dfBaronies.loc[dfBaronies['Cell ID'] == cellId, 'Kingdom'] = kingdomName
            #         countMapCounties[countyName] += 1
            #         # del countMapCounties[c]
            #         moved = True
            #         break
            #     if moved == False:
            #         print("Is this an island?")
            #         print(cellId, countyName, kingdomName, "nb:", nb)
                    
            # print(neighbors)
    print(dfBaronies)
    dfBaronies.to_excel(duchyDef_file, index=False)
    return

    # ############## Will be extended when working ############    
    # kingdoms = dfBaronies["Kingdom"].unique()
    # # print(kingdoms)
    # # print(type(kingdoms))
    # countMapKingdoms = dict()
    # for k in kingdoms:
    #     baronyList = dfBaronies.loc[dfBaronies['Kingdom'] == k, 'Cell ID'].tolist()
    #     countMapKingdoms[k] = len(baronyList)
    # # print(countMapKingdoms)
    # countMapKingdoms = sorted(countMapKingdoms.items(), key=lambda x: x[1])
    # # countMapKingdoms = dict(sorted(countMapKingdoms.items(), key=lambda item: item[1]))
    # # print(countMapKingdoms)
    # for k in countMapKingdoms:
    #     # print(k)
    #     # print(type(k))
    #     # print(k[1])
    #     kName = k[0]
    #     kCount = k[1]
    #     if kCount < 5:
    #         print(kName, kCount)
    #         # TODO: this should be included into another kingdom
    #         # e.g. look who is neighbor
    #     elif kCount < 10:
    #         print(kName, kCount)
    #         # TODO
    #     # cells_data['neighbors'] = cells_data['neighbors'].apply(lambda x: [int(n) for n in x.strip('[]').split(',')])
    # # geht sowas?
    # # grouped_df = dfBaronies.groupby('Kingdom').agg('Cell ID').count()
    # dfBaronies.to_excel(duchyDef_file, index=False)
    
def finalorder(duchyDef_file, final_file):
    # Unnamed: 0	Unnamed: 1	R	G	B	Nearest Town	Cell ID	population	Unnamed: 8	Kingdom	Unnamed: 10	County	Culture	Religion
    # Load Excel file into a pandas DataFrame    
    df = pd.read_excel(duchyDef_file, usecols=['0', 'ExcelIdx', 'R', 'G', 'B', 'Nearest Town', 'Cell ID', 'population', 'Empire', 'Kingdom', 'Duchy', 'County', 'Culture', 'Religion' ])
    print(df)
    # Check if columns G and H exist by number
    if len(df.columns) > 7:
        # Delete values in columns G and H
        df.iloc[:, [6, 7]] = ""
    # Check if columns J and L exist by number
    if len(df.columns) > 9:
        # Copy values from Column J to Column I
        df.iloc[:, 8] = df.iloc[:, 9]
        # Copy values from Column L to Column K
        df.iloc[:, 10] = df.iloc[:, 11]
    # Save modified DataFrame back to Excel file
    df.to_excel(final_file, index=False)



biomes_file = "../ck3_map_gen_data/_snibmap/biomes.xlsx"    # not changed

outputDir = "../ck3_map_gen_data/_snibmap/"
updated_file = "../ck3_map_gen_data/_snibmap/updated_file.xlsx"
combined_data_file = "../ck3_map_gen_data/_snibmap/combined_data_2.xlsx"
cells_data_file = "../ck3_map_gen_data/_snibmap/cellsData.xlsx"
bfs_distance_file = "../ck3_map_gen_data/_snibmap/bfs_distance_file.xlsx"
bfs_distance_file_with_neighbors = "../ck3_map_gen_data/_snibmap/bfs_distance_file_with_neighbors.xlsx"
output_file_extended_neighbors = "../ck3_map_gen_data/_snibmap/snibboOutputTest_newBaronies_extended.xlsx"
output_png = "../ck3_map_gen_data/_snibmap/provinces.png"
provinceDef_file = "../ck3_map_gen_data/_snibmap/provinceDef.xlsx"
provinceDef_file_transformed = "../ck3_map_gen_data/_snibmap/provinceDefTrans.xlsx"
duchyDef_file = "../ck3_map_gen_data/_snibmap/duchyDef.xlsx"
final_file = "../ck3_map_gen_data/_snibmap/finalBFS.xlsx"

# maxSizeBarony = 6      # warning 6 might take a long time... testing
maxSizeBarony = 20      # 20 is a good measure
# maxSizeBarony = 2     # should be tested as well
useCounties = False

start_time = time.time()
print(time.ctime())
# bfs_distance_to_new_baronies(combined_data_file, cells_data_file, maxSizeBarony, useCounties)    # new
# bfs_distance(combined_data_file, cells_data_file, bfs_distance_file, useCounties)  # adjusted
# bfs_appendneighbors(cells_data_file, bfs_distance_file, bfs_distance_file_with_neighbors)  # new

# colorRandomBFS(bfs_distance_file_with_neighbors)     # adjusted
# provinceMapBFS(bfs_distance_file_with_neighbors, 150, output_png)    # nothing changed

# extractBFS(bfs_distance_file_with_neighbors, provinceDef_file)   # nothing changed
# BaronyId(combined_data_file, provinceDef_file)  # nothing changed
# bfs_calcType(biomes_file, cells_data_file, provinceDef_file)      # not finished yet

# bfs_putNeighborsToTowns(cells_data_file, bfs_distance_file_with_neighbors, provinceDef_file, provinceDef_file_transformed)

# ProvData(updated_file, provinceDef_file_transformed, provinceDef_file_transformed)  # nothing changed
# cOrder(provinceDef_file_transformed)    # adjusted

# assignEmpireAndDuchyId(provinceDef_file_transformed, duchyDef_file)

# TODO: maybe we have to delete columns to fit with INTERFACE to map gen
# TODO!!
# finalorder(duchyDef_file, final_file)


end_time = time.time()
print(time.ctime())
print("--- %s seconds ---" % (round(end_time - start_time, 2)))

# # bfs_fillneighbors(combined_data_file, output_file, output_file_extended_neighbors)    # new >> not required?


# def bfs_fillNeighbors(combined_data_file, input_file, output_file):
#     # # neighbors needs to be "two sided" - with the above it might be possible that a neighbour will exists only after a new frontier was set
#     # neighbors = dict()
#     # # Read current data
#     # output_data = pd.read_excel(input_file, sheet_name="Sheet1")
#     # # Loop over Ids
#     # for id in output_data["id"]:
#     #     if output_data.loc[output_data["id"] == id, "distance"].tolist()[0] == 0:
#     #         # Get neighbors and loop them
#     #         nb = output_data.loc[output_data["id"] == id, "neighbors"].tolist()[0]
#     #         xList = ast.literal_eval(nb)
#     #         neighbors[id] = xList
#     #         for x in xList:
#     #             # Get neighbors from neighbour:
#     #             checker = output_data.loc[output_data["id"] == x, "neighbors"].tolist()
#     #             if len(checker) != 0:       # always true, when using "bfs_distance_to_new_baronies"
#     #                 nn = output_data.loc[output_data["id"] == x, "neighbors"].tolist()[0]
#     #                 nbs = ast.literal_eval(nn)
#     #                 if len(nbs) == 0:
#     #                     neighbors[x] = []
#     #                 else:
#     #                     neighbors[x] = nbs
#     #                 if id not in nbs:               
#     #                     neighbors[x].append(id)
#     # output_data["neighbors_new"] = output_data["id"].apply(lambda x: neighbors.get(x))
#     # output_data.to_excel(output_file, index=False)
    
#     # neighbors.clear()
#     # towns = output_data["nearest_town"].unique()
#     # for t in towns:
#     #     nb = output_data.loc[output_data["nearest_town"] == t, "neighbors_new"].tolist()
#     #     newList = []
#     #     checker = output_data.loc[output_data["id"] == x, "neighbors"].tolist()
#     #     if len(checker) != 0:       # always true, when using "bfs_distance_to_new_baronies"
#     #         for n in nb:
#     #             for x in ast.literal_eval(n):                
#     #                 if x not in newList:
#     #                     newList.append(x)
#     #     neighbors[t] = newList
#     # output_data["neighbors"] = output_data["nearest_town"].apply(lambda x: neighbors.get(x))
#     # output_data.to_excel(output_file, index=False)

#     # neighbors needs to be "two sided" - with the above it might be possible that a neighbour will exists only after a new frontier was set
#     neighbors = dict()            
#     # Read current data
#     output_data = pd.read_excel(input_file, sheet_name="Sheet1")    
#     towns = output_data["nearest_town"].unique()
#     for town in towns:
#         newList = []
#         townlist_neighbors = output_data.loc[output_data["nearest_town"] == town, "neighbors"].tolist()
#         for singleItem in townlist_neighbors:
#             for x in ast.literal_eval(singleItem):
#                 if x not in newList:
#                     newList.append(x)
#         neighbors[town] = newList

#     # Create a dictionary to map cell IDs to town names
#     town_data = pd.read_excel(combined_data_file, sheet_name="burgs")
#     town_to_id = dict(zip(town_data["name"], town_data["cell"]))    

#     # Loop over Towns    
#     for t in neighbors:
#         nb = neighbors[t]
#         for id in nb:
#             nb_nbs = output_data.loc[output_data["id"] == id, "neighbors"].tolist()[0]
#             nbList = ast.literal_eval(nb_nbs)
#             if town_to_id[t] not in nbList:
#                 # add town_to_id[t] to output_data.loc[output_data["id"] == id, "neighbors"]
#                 currentData = output_data.loc[output_data["id"] == id, "neighbors"].tolist()[0]
#                 nbs = ast.literal_eval(currentData)
#                 nbs.append(id)
#                 output_data.loc[output_data["id"] == id, "neighbors"].value = nbs
#     output_data.to_excel(output_file, index=False)