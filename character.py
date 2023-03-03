import os
import pandas as pd
import random
import re
import pandas as pd
import os
import random
from openpyxl import Workbook
from random import choice


#Generates Birthday from Bookmark
def generate_random_date():
    BM = 1066
    maxyear = BM - 100
    minyear = BM - 5
    year = random.randint(maxyear, minyear)
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    return f"{year}.{month}.{day}"


#WIP Generates Death date from Birthday
def generate_random_ddate(bdate):
    random_days = int(random.normalvariate(70, 10))
    year, month, day = bdate
    days_in_month = 28 + (month + month//8) % 2 + 2 % month + 2 * (1 // month)
    day += random_days
    if day > days_in_month:
        day -= days_in_month
        month += 1
        if month > 12:
            month = 1
            year += 1
    return f"{year}.{month}.{day}"



#Generates Rulers to Excel file - Was originally generating to text but excel better for project long term
def rulerGenXL(combined_data, output):
    print(combined_data)
    print(output)
    # Read in the combined_data.xlsx file and the "states" sheet
    data = pd.read_excel(combined_data, sheet_name="states")

    # Read in the cultures data
    cultures_data = pd.read_excel(combined_data, sheet_name="cultures")

    # Read in the religions data
    religion_data = pd.read_excel(combined_data, sheet_name="religion")

    # Initialize a dictionary to keep track of the unique names
    unique_names = {}

    # Initialize an empty list to hold the character data
    character_data = []
    states = 2

    # Loop through the rows of the data
    for index, row in data.iterrows():
        # Get the name from the "name" column
        name = row["name"]

        # Generate a unique ID number
        if name not in unique_names:
            unique_names[name] = len(unique_names) + 1


        #Discovered this method of having a random few states not appear at game start by assigning the character an invalid ID. Figured it could be worked into a feature for user input to determine how many states should exist at game start
        if states == 1:
            #Generates all -9 states since 9 ID's begin with 0 which is invalid
            id_num = f"{unique_names[name]:02d}0001"
        else:
            #Generates all states
            id_num = f"9{unique_names[name]:02d}001"



        culture = choice(cultures_data["name"].tolist())
        religion = choice(religion_data["name"].tolist())

        # Create a dictionary with the character's information
        character_info = {}
        character_info["id_num"] = id_num
        character_info["name"] = name
        character_info["bdate"] = generate_random_date()
        character_info["culture"] = culture
        character_info["religion"] = religion
        character_info["charname"] = random.choice(["Abarat", "Adamar", "Adorellan", "Adresin", "Aelrindel", "Aerendyl", "Aeson", "Afamrail",
             "Agandaur", "Agis", "Aias", "Aiduin", "Aien", "Ailas", "Ailduin", "Ailen", "Ailluin", "Ailmar",
             "Ailmer", "Ailmon", "Ailre", "Ailred", "Ailuin", "Ailwin", "Aimar", "Aimer", "Aimon", "Airdan",
             "Aire", "Aired", "Aithlin", "Aiwin", "Akkar", "Alabyran", "Alaion", "Alas", "Alen", "Alinar",
             "Alluin", "Almar", "Almer", "Almon", "Alok", "Alosrin", "Alre", "Alred", "Althidon", "Alwin",
             "Amrynn", "Andrathath", "Anfalen", "Anlyth", "Aolis", "Aquilan", "Arathorn", "Arbane",
             "Arbelladon", "Ardreth", "Ardryll", "Arel", "Arlen", "Arun", "Ascal", "Athtar", "Aubron",
             "Aumanas", "Aumrauth", "Avourel", "Ayas", "Ayduin", "Ayen", "Ayluin", "Aymar", "Aymer", "Aymon",
             "Ayre", "Ayred", "Aywin", "Belanor", "Beldroth", "Bellas", "Beluar", "Biafyndar", "Bialaer",
             "Braern", "Cailu", "Camus", "Castien", "Chathanglas", "Cohnal", "Conall", "Connak", "Cornaith",
             "Corym", "Cyran", "Dain", "Dakath", "Dalyor", "Darcassan", "Darfin", "Darthoridan", "Darunia",
             "Deldrach", "Delmuth", "Delsaran", "Devdan", "Drannor", "Druindar", "Durlan", "Durothil",
             "Dyffros", "Edwyrd", "Edyrm", "Ehlark", "Ehrendil", "Eilauver", "Elaith", "Elandorr", "Elas",
             "Elashor", "Elauthin", "Eldaerenth", "Eldar", "Eldrin", "Elduin", "Elen", "Elephon", "Elidyr",
             "Elion", "Elkhazel", "Ellisar", "Elluin", "Elmar", "Elmer", "Elmon", "Elnaril", "Elorshin",
             "Elpharae", "Elre", "Elred", "Eltaor", "Elwin", "Elyon", "Emmyth", "Entrydal", "Erendriel",
             "Eriladar", "Erlan", "Erlareo", "Erlathan", "Eroan", "Erolith", "Estelar", "Ethlando", "Ettrian",
             "Evindal", "Faelar", "Faelyn", "Faeranduil", "Falael", "Felaern", "Fenian", "Feno", "Feyrith",
             "Fhaornik", "Filarion", "Filvendor", "Filverel", "Flardryn", "Flinar", "Folas", "Folduin",
             "Folen", "Folluin", "Folmar", "Folmer", "Folmon", "Folre", "Folred", "Folwin", "Fylson",
             "Gaeleath", "Gaelin", "Galaeron", "Galan", "Galather", "Ganamede", "Gantar", "Garrik",
             "Garynnon", "Giullis", "Glanduil", "Glarald", "Glorandal", "Goras", "Gorduin", "Goren",
             "Gorluin", "Gormar", "Gormer", "Gormon", "Gorre", "Gorred", "Gorwin", "Grathgor", "Haemir",
             "Hagas", "Hagduin", "Hagen", "Hagluin", "Hagmar", "Hagmer", "Hagre", "Hagred", "Hagwin",
             "Haladavar", "Halafarin", "Halamar", "Haldir", "Halflar", "Halueth", "Halueve", "Hamon", "Haryk",
             "Hastios", "Hatharal", "Horith", "Hubys", "Iefyr", "Ievis", "Ilbryen", "Ilimitar", "Iliphar",
             "Illianaro", "Illithor", "Illitran", "Ilphas", "Ilrune", "Ilthuryn", "Ilvisar", "Inchel",
             "Inialos", "Intevar", "Iolas", "Iolrath", "Itham", "Ivaran", "Ivasaar", "Iymbryl", "Iyrandrar",
             "Jandar", "Jannalor", "Jaonos", "Jassin", "Jhaan", "Jhaartael", "Jhaeros", "Jonik", "Jorildyn",
             "Kailu", "Katar", "Katyr", "Kellam", "Kelvhan", "Kendel", "Kerym", "Keryth", "Kesefeon",
             "Kharis", "Khatar", "Khidell", "Khiiral", "Khilseith", "Khuumal", "Khyrmin", "Kieran", "Kiirion",
             "Kindroth", "Kivessin", "Klaern", "Kolvar", "Kuskyn", "Kymil", "Kyrenic", "Kyrtaar", "Laeroth",
             "Lafarallin", "Laiex", "Lamruil", "Larongar", "Larrel", "Lathai", "Lathlaeril", "Lhoris",
             "Lianthorn", "Llarm", "Llewel", "Lorsan", "Luirlan", "Luthais", "Luvon", "Lyari", "Lyklor",
             "Lysanthir", "Maeral", "Maiele", "Malgath", "Malon", "Maradeim", "Marikoth", "Marlevaur",
             "Melandrach", "Merellien", "Merith", "Methild", "Mhaenal", "Mitalar", "Mihangyl", "Miirphys",
             "Mirthal", "Mlartlar", "Mnementh", "Morthil", "Myrdin", "Myriil", "Myrin", "Mythanar", "Naertho",
             "Naeryndam", "Naesala", "Narbeth", "Nardual", "Nasir", "Navarre", "Nelaeryn", "Neldor",
             "Neremyn", "Nesterin", "Nevarth", "Nhamashal", "Nieven", "Nindrol", "Ninleyn", "Ninthalor",
             "Niossae", "Nuvian", "Nylian", "Nym", "Nyvorlas", "Olaurae", "Onas", "Oncith", "Onvyr", "Orist",
             "Ornthalas", "Orrian", "Orym", "Otaehryn", "Othorion", "Paeral", "Paeris", "Pelleas", "Phaendar",
             "Pharom", "Phraan", "Pirphal", "Purtham", "Pyrravyn", "Pywaln", "Qildor", "Raeran", "Raibyn",
             "Ralnor", "Ranaeril", "Rathal", "Reluraun", "Reluvethel", "Rennyn", "Reptar", "Respen",
             "Revalor", "Rhalyf", "Rhangyl", "Rhistel", "Rhothomir", "Rhys", "Rilitar", "Riluaneth", "Rolim",
             "Rothilion", "Ruehnar", "Ruith", "Ruvaen", "Ruven", "Ruvyn", "Rychell", "Rydel", "Ryfon", "Ryo",
             "Ryul", "Saelethil", "Saevel", "Saleh", "Samblar", "Sanev", "Scalanis", "Selanar", "Sharian",
             "Shaundyl", "Shyrrik", "Sihnion", "Silvyr", "Simimar", "Sinaht", "Siveril", "Sontar", "Sudryal",
             "Sundamar", "Sylvar", "Sythaeryn", "Taegen", "Taenaran", "Taeral", "Taerentym", "Taleasin",
             "Tamnaeth", "Tanithil", "Tannatar", "Tannivh", "Tannyll", "Tanyl", "Tanyth", "Taranath",
             "Tarathiel", "Taredd", "Tarron", "Tasar", "Tassarion", "Tathaln", "Thalanil", "Thallan",
             "Theodas", "Theodemar", "Theoden", "Theodluin", "Theodmer", "Theodmon", "Theodre", "Theodred",
             "Thuridan", "Tiarsus", "Tolith", "Tordynnar", "Toross", "Traeliorn", "Travaran", "Triandal",
             "Ualiar", "Uevareth", "Uldreyin", "Urdusin", "Usunaar", "Uthorim", "Vaalyun", "Vaeril", "Vamir",
             "Varitan", "Velethuil", "Venali", "Vesryn", "Vesstan", "Virion", "Volodar", "Voron", "Vuduin",
             "Vulas", "Vulen", "Vulluin", "Vulmar", "Vulmer", "Vulmon", "Vulre", "Vulred", "Vulwin",
             "Wirenth", "Wistari", "Wyn", "Wyninn", "Wyrran", "Yalathanil", "Yesanith", "Yhendorn", "Ylyndar",
             "Zaos", "Zelphar", "Zeno", "Zhoron"])

        # Add the character's information to the list
        character_data.append(character_info)

    # Create a new workbook and sheet to hold the character data
    wb = Workbook()
    ws = wb.active
    ws.append(["id_num", "name", "bdate", "culture", "religion","charname"])


    # Write the character data to the sheet
    for character in character_data:
        ws.append([
            character["id_num"],
            character["name"],
            character["bdate"],
            character["culture"],
            character["religion"],
            character["charname"]

        ])

    # Save the workbook to a file
    #filename = "characters.xlsx"
    wb.save(output)


#Writes characters to files, from the spreadsheet generated previous
def rulerWrite(input,output):
    # Read the excel file and extract unique names
    df = pd.read_excel(input)
    unique_names = df["name"].unique()

    # Create subfolder if it doesn't exist
    subfolder = output
    os.makedirs(output, exist_ok=True)

    # Iterate through unique names and create text files
    for name in unique_names:
        filename = os.path.join(output, f"{name}.txt")
        with open(filename, "w") as f:
            # Write the character's information to the file
            rows = df.loc[df["name"] == name]
            for i, row in rows.iterrows():
                id_num = row["id_num"]
                birth_date = row["bdate"]
                religion = row["religion"]
                culture = row["culture"]
                charname = row["charname"]

                #process inputs to match id's
                relproc = re.sub(r'\W+', '', religion).lower()
                culproc = re.sub(r'\W+', '', culture).lower()

                f.write(f"{id_num} = {{\n")
                f.write(f"\tname = {charname}\n")
                f.write(f"\treligion = {relproc}\n")
                f.write(f"\tculture = {culture}\n")
                f.write(f"\t{birth_date} = {{\n")
                f.write(f"\t\tbirth = yes\n")
                f.write(f"\t}}\n")
                f.write("}")


#Modifies Title files generated by Map Filler to assign characters their titles
def modHistory(characters_file, titles_folder):
    import os
    import openpyxl
    # Read the character information from the Excel file
    workbook = openpyxl.load_workbook(characters_file)
    worksheet = workbook.active

    characters = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        id_num, name, bdate = row[0], row[1], row[2]
        characters[name.lower()] = {"id_num": id_num, "bdate": bdate}

    # Iterate over the text files in the specified directory
    for filename in os.listdir(titles_folder):
        if filename.endswith(".txt"):

            filepath = os.path.join(titles_folder, filename)
            #print("Editing ", filepath)
            # Read the contents of the file
            with open(filepath, "r") as f:
                contents = f.read()

            # Modify any instances of k_{name} = { 800.1.1 = { liege = e_{name} } }
            for name, info in characters.items():
                k_name = "k_" + name.lower()
                k_name_upper = k_name.upper()
                id_num, bdate = info["id_num"], info["bdate"]
                new_line = f"k_{name} = {{\n    800.1.1 = {{\n        liege = e_{name}\n    }}\n    {bdate} = {{\n        holder = {id_num}\n    }}\n}}"



                if f"{k_name} = {{\n\t800.1.1={{\n\t\tliege=e_{name}\n\t}}\n}}" in contents:

                    print(f"Generating Ruler for {k_name} in {filepath}")
                    contents = contents.replace(f"{k_name} = {{\n\t800.1.1={{\n\t\tliege=e_{name}\n\t}}\n}}", new_line)

            # Write the modified contents back to the file
            with open(filepath, "w") as f:
                #print(f"Writing to file: {filepath}")
                f.write(contents)


#modHistory()



